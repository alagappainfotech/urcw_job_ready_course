"""
Module 7: Data Handling, Persistence & Exploration - Exercises and Labs
"""

from pathlib import Path
import csv
import json
import sqlite3
import requests
import xmltodict
from openpyxl import Workbook, load_workbook

try:
    import pandas as pd
    import matplotlib.pyplot as plt
except Exception:
    pd = None
    plt = None


# Quick Checks

def qc_stdlib_for_tasks():
    tasks = [
        ("Move a file", "shutil.move"),
        ("Walk directories", "pathlib.Path.rglob / os.walk"),
        ("Parse CSV", "csv"),
        ("Parse JSON", "json"),
        ("Fetch HTTP JSON", "requests.get().json()"),
        ("SQLite DB", "sqlite3"),
    ]
    for desc, mod in tasks:
        print(f"- {desc}: {mod}")


# Try This

def try_read_pipe_delimited(text_path: str = "./pipe_data.txt"):
    p = Path(text_path)
    if not p.exists():
        p.write_text("name|age|score\nAlice|25|91\nBob|30|88\n")
    rows = []
    for line in p.read_text().splitlines()[1:]:
        name, age, score = [x.strip() for x in line.split("|")]
        rows.append({"name": name, "age": int(age), "score": float(score)})
    print("Rows:", rows)


def try_fetch_and_parse_json(url: str = "https://httpbin.org/json"):
    resp = requests.get(url, timeout=10)
    resp.raise_for_status()
    data = resp.json()
    print("Top-level keys:", list(data.keys()))


# Labs

def lab_weather_api_transform(url: str = "https://api.met.no/weatherapi/locationforecast/2.0/compact?lat=59.9&lon=10.7"):
    """Fetch JSON weather, transform to flat records list (date, temp)."""
    headers = {"User-Agent": "educational-script"}
    resp = requests.get(url, headers=headers, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    timeseries = data.get("properties", {}).get("timeseries", [])
    records = []
    for entry in timeseries[:24]:  # first 24 entries
        time = entry.get("time")
        temp = entry.get("data", {}).get("instant", {}).get("details", {}).get("air_temperature")
        if time is not None and temp is not None:
            records.append({"time": time, "temperature_c": temp})
    print("First 5 records:", records[:5])
    return records


def lab_sqlite_create_and_query(db_path: str = "./module7.db"):
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS sales (id INTEGER PRIMARY KEY, name TEXT, amount REAL, month TEXT)")
    cur.executemany(
        "INSERT INTO sales(name, amount, month) VALUES(?,?,?)",
        [("Alice", 120.5, "2024-01"), ("Bob", 80.0, "2024-01"), ("Alice", 99.9, "2024-02")],
    )
    conn.commit()
    cur.execute("SELECT name, SUM(amount) FROM sales WHERE month=? GROUP BY name", ("2024-01",))
    print("Totals in 2024-01:", cur.fetchall())
    conn.close()


def try_excel_read_write(path: str = "./demo.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "age"]) 
    ws.append(["Alice", 25])
    ws.append(["Bob", 30])
    wb.save(path)
    wb2 = load_workbook(path)
    ws2 = wb2.active
    rows = [[c.value for c in row] for row in ws2.iter_rows(min_row=1, max_col=2, max_row=3)]
    print("Excel rows:", rows)


def try_xml_fetch_and_parse(url: str = "https://www.w3schools.com/xml/simple.xml"):
    resp = requests.get(url, timeout=10)
    resp.raise_for_status()
    data = xmltodict.parse(resp.text)
    print("XML root keys:", list(data.keys()))


def try_pandas_groupby_plot():
    if pd is None:
        print("pandas/matplotlib not available")
        return
    df = pd.DataFrame({
        "name": ["Alice", "Alice", "Bob", "Bob"],
        "month": ["2024-01", "2024-02", "2024-01", "2024-02"],
        "amount": [120.5, 99.9, 80.0, 110.0]
    })
    agg = df.groupby(["name", "month"])['amount'].sum().reset_index()
    print("Aggregated:\n", agg)
    if plt is not None:
        pivot = agg.pivot(index="month", columns="name", values="amount")
        pivot.plot(kind="line", marker="o")
        plt.title("Monthly totals by person")
        plt.tight_layout()
        plt.savefig("module7_plot.png")
        print("Saved plot to module7_plot.png")


def try_sentinel_stop_on_dash(path: str = "./sentinel.txt"):
    p = Path(path)
    if not p.exists():
        p.write_text("A\nB\n---\nC\n")
    out = []
    for line in p.read_text().splitlines():
        if line.strip() == "---":
            break
        out.append(line)
    print("Before sentinel:", out)


def qc_nosql_prompts():
    print("Redis: caching, session storage, rate limits; not for complex relational queries.")
    print("MongoDB: flexible document storage; schema evolution; not ideal for multi-document transactions.")


if __name__ == "__main__":
    qc_stdlib_for_tasks()
    try_read_pipe_delimited()
    try_fetch_and_parse_json()
    lab_weather_api_transform()
    lab_sqlite_create_and_query()
    try_excel_read_write()
    try_xml_fetch_and_parse()
    try_pandas_groupby_plot()
    try_sentinel_stop_on_dash()
    qc_nosql_prompts()


